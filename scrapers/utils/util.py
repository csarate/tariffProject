import datetime
import time
import os
import openpyxl
import json
import unicodedata
import re
import boto3
from io import BytesIO
from botocore.errorfactory import ClientError
log_file_path='\\Log/log_scraping.xlsx'

def find_missing(lst): # this funtion is used to find missing values in a list of numbers
    return [x for x in range(int(lst[0]), int(lst[-1])+1)
                               if str(x) not in lst]

def delete_file_s3 (filepath): # this function deletes a file from a S3 bucket
    s3_resource = boto3.resource('s3')
    current_dir=os.getcwd().rsplit("\\",1)[0]
    with open( os.path.join(current_dir,"bucketname.txt") ,"r") as f:
          bucket_link= f.read()
    s3_resource.Object(bucket_link, filepath).delete()

def create_console_log(): # this function creates the console_log case necessary
    try:
        current_dir = os.getcwd().rsplit("\\", 1)[0]
        with open(os.path.join(current_dir, "bucketname.txt"), "r") as f:
            bucket_link = f.read()
        s3_client = boto3.client('s3')
        s3_client.get_object(Bucket=bucket_link, Key="log/console-log.json")
    except ClientError:
        dump_files_s3([], "log/console-log.json")
        log("Created console log in S3")

def update_console_log(new_info): # this updates the console_log.json with new info
    current_dir = os.getcwd().rsplit("\\", 1)[0]
    with open(os.path.join(current_dir, "bucketname.txt"), "r") as f:
        bucket_link = f.read()
    s3_client = boto3.client('s3')
    try:
        response = s3_client.get_object(Bucket=bucket_link, Key="log/console-log.json")
        if response is not None:
            content = json.loads(response['Body'].read().decode('utf-8'))
            content.append([new_info])
            try:
                response = s3_client.put_object(Bucket=bucket_link, Body=(bytes(json.dumps(content).encode('UTF-8'))),
                                                Key="log/console-log.json")
                return response
            except Exception as e:
                time.sleep(1)
                log("Couldn't save log/console-log.json information in Amazon S3. Error description:" + str(e))
        else:
            log("Couldn't load log/console-log.json information in Amazon S3. Error description:" + response)
            time.sleep(1)
            return []
    except Exception as e:
        time.sleep(1)
        log("Couldn't load log/console-log.json information in Amazon S3. Error description:" + str(e))



def create_already_download_files(filename): # create a json file to store informations of the gathered links of the companies
    try:
        current_dir = os.getcwd().rsplit("\\", 1)[0]
        with open(os.path.join(current_dir, "bucketname.txt"), "r") as f:
            bucket_link = f.read()
        s3_client = boto3.client('s3')
        s3_client.get_object(Bucket=bucket_link, Key="to_be_downloaded/" + filename)
    except ClientError:
        dump_files_s3([], "to_be_downloaded/"+filename)
        log("Created "+ filename + " in S3")

def bucket_contents(folder_path): # check the contents of a bucket
    current_dir=os.getcwd().rsplit("\\",1)[0]
    with open( os.path.join(current_dir,"bucketname.txt") ,"r") as f:
          bucket_link= f.read()
    s3_client = boto3.resource('s3')
    my_bucket= s3_client.Bucket(bucket_link)
    content_list = []
    for object_summary in my_bucket.objects.filter(Prefix=folder_path):
        content_list.append(object_summary.key)
    return content_list

def search_parts(filename, folder_path): # search the bucket for file parts
    files_list = bucket_contents(folder_path)
    numbers_list = []
    for item in files_list:
        try:
            item = item.split("/")[-1]
            file_key = item.rsplit(".",1)[0].rsplit("_",1)[0]
            file_number = item.rsplit(".")[0].split("_")[-1]
            file_extension = item.split(".")[-1]
            file_to_compare_key = filename.rsplit(".")[0]
            file_to_compare_extension = filename.split(".")[-1]
        except:
            continue
        if file_key == file_to_compare_key and file_extension == file_to_compare_extension:
            numbers_list.append(file_number)
    try:
        missing_numbers = find_missing(numbers_list)
    except:
        return 0
    if len(numbers_list) == 0:
        return 0
    if len(missing_numbers) == 0:
        return max(numbers_list)
    else:
        return min(missing_numbers)

def save_directly_s3(file, filename, outputpath):
    try:
        s3_client = boto3.client('s3')
        current_dir = os.getcwd().rsplit("\\" , 1)[0]
        with open(os.path.join(current_dir , "bucketname.txt") , "r") as f:
            bucket_link = f.read()
        file_to_save=BytesIO(file.read())
        file_to_save = file_to_save.getvalue()
        response = s3_client.put_object(Bucket=bucket_link,
                                        Body=(bytes(file_to_save)),
                                        Key=outputpath+"/"+filename)
        return response
    except Exception as e:
        log("Couldn't save"+ filename +" information in Amazon S3. Error description:" + str(e))

def slugify(value, allow_unicode=False): # standarize the name of the files
    """
    Taken from https://github.com/django/django/blob/master/django/utils/text.py
    """
    value = str(value)
    if allow_unicode:
        value = unicodedata.normalize('NFKC', value)
    else:
        value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    value = re.sub(r'[^\w\s-]', '', value.lower())
    return re.sub(r'[-\s]+', '-', value).strip('-_')
def log(*args, end="\n"):
    dt = datetime.datetime.now().strftime("[%H:%M:%S] ->")
    print("{}".format(dt), " ".join([str(i) for i in args]), end=end)

def crdownload_remover(path_to_downloads): # removes uncomplete download files from path
    for fname in os.listdir(path_to_downloads):
        if fname.endswith('.crdownload'):
            os.remove(os.path.join(path_to_downloads, fname))

def download_wait(path_to_downloads): # wait download to be finished
    seconds = 0
    dl_wait = True
    while dl_wait and seconds < 360:
        time.sleep(1)
        dl_wait = False
        for fname in os.listdir(path_to_downloads):
            if fname.endswith('.crdownload'):
                dl_wait = True
        seconds += 1
    return seconds

def create_day_in_log_file(sheet_name): # creates a day in a specific country sheet the log_scraping.xlsx
    workbook = read_log_file_s3()
    today_string = datetime.datetime.today().strftime("%d-%m-%Y")
    log_wb = openpyxl.load_workbook(filename=BytesIO(workbook))
    enlaces_sh = log_wb[sheet_name]
    col_count_scraper = 0
    while enlaces_sh.cell(2, 2 + col_count_scraper).value != None or enlaces_sh.cell(2, 2 + col_count_scraper).value=="OK":
        col_count_scraper += 1
    need_to_create_scraper = True
    for i in range(col_count_scraper):
        if today_string in enlaces_sh.cell(2, 2 + i).value:
            log('Date already created in log file for scraper.')
            need_to_create_scraper = False
            break
    if need_to_create_scraper is True:
        enlaces_sh.cell(2, 2 + col_count_scraper, today_string)
    save_log_file_s3(log_wb)
    log_wb.close()


def register_positive_on_log(sheet_name, row): # register positive status of scraper for a company in log_scraping.xlsx
    workbook = read_log_file_s3()
    today_string = datetime.datetime.today().strftime("%d-%m-%Y")
    log_wb = openpyxl.load_workbook(filename=BytesIO(workbook))
    enlaces_sh = log_wb[sheet_name]
    col_count = 0
    while enlaces_sh.cell(2, 2 + col_count).value != today_string:
        col_count += 1
    enlaces_sh.cell(row+2, 2 + col_count, 'OK')
    save_log_file_s3(log_wb)
    log_wb.close()

def register_negative_on_log(sheet_name, row): #register negative status of scraper for a company in log_scraping.xlsx
    workbook=read_log_file_s3()
    today_string = datetime.datetime.today().strftime("%d-%m-%Y")
    log_wb = openpyxl.load_workbook(filename=BytesIO(workbook))
    enlaces_sh = log_wb[sheet_name]
    col_count = 0
    while enlaces_sh.cell(2, 2 + col_count).value != today_string:
        col_count += 1
    enlaces_sh.cell(row, 2 + col_count, 'FAIL')
    log_wb.save(log_wb)
    log_wb.close()

def remove_repeated_files(download_dir): # remove repeated files from local dir. from download dir DEPRECATED
    all_downloaded_files = os.listdir(download_dir)
    for file in all_downloaded_files:
        if '(1)' in file:
            os.remove(download_dir + '/' + file)

def dump_files(data_list,filepath): # dump json file in local dir. DEPRECATED
    with open(filepath, 'w') as outfile:
        json.dump(data_list, outfile)
    outfile.close()

def dump_files_s3(data_list,filepath): # dump list in S3 given a list data and the filepath in S3 Bucket
    current_dir=os.getcwd().rsplit("\\",1)[0]
    with open( os.path.join(current_dir,"bucketname.txt") ,"r") as f:
          bucket_link= f.read()
    s3_client = boto3.client('s3')
    try:
        response = s3_client.put_object(Bucket=bucket_link,
        Body=(bytes(json.dumps(data_list).encode('UTF-8'))),
        Key=filepath)
        return response
    except Exception as e:
        log("Couldn't save " + filepath + " information in Amazon S3. Error description:" + str(e))

def dump_files_s3_v2(data_list,filepath): # Dump bytes information in S3
    current_dir=os.getcwd().rsplit("\\",1)[0]
    with open( os.path.join(current_dir,"bucketname.txt") ,"r") as f:
          bucket_link= f.read()
    s3_client = boto3.client('s3')
    try:
        response = s3_client.put_object(Bucket=bucket_link,
        Body=(bytes(data_list)),
        Key=filepath)
        return response
    except Exception as e:
        log("Couldn't save " + filepath + " information in Amazon S3. Error description:" + str(e))


def read_files_s3(filepath): # read json from S3
    current_dir=os.getcwd().rsplit("\\",1)[0]
    with open( os.path.join(current_dir,"bucketname.txt") ,"r") as f:
          bucket_link= f.read()
    s3_client = boto3.client('s3')
    try:
        response = s3_client.get_object(Bucket=bucket_link, Key=filepath)
        if response is not None:
            content= json.loads(response['Body'].read().decode('utf-8'))
            return content
        else:
            log("Couldn't load " + filepath + " information in Amazon S3. Error description:" + response)
            return []
    except Exception as e:
        log("Couldn't load " + filepath + " information in Amazon S3. Error description:" + str(e))

def read_log_file_s3(): # read log_scraping.xlsx from S3
    filepath="log/log_scraping.xlsx"
    s3_client = boto3.client('s3')
    current_dir=os.getcwd().rsplit("\\",1)[0]
    with open( os.path.join(current_dir,"bucketname.txt") ,"r") as f:
          bucket_link= f.read()
    try:
        response = s3_client.get_object(Bucket=bucket_link, Key=filepath)
        response=response['Body'].read()
        if response is not None:
            return response
        else:
            log("Couldn't load " + filepath + " information in Amazon S3. Error description:" + response)
            return []
    except Exception as e:
        log("Couldn't load " + filepath + " information in Amazon S3. Error description:" + str(e))

def save_log_file_s3(file):# save log_scraping.xlsx and log_scraping_backup.xlsx from S3
    try:
        current_dir = os.getcwd().rsplit("\\", 1)[0]
        with open(os.path.join(current_dir, "bucketname.txt"), "r") as f:
            bucket_link = f.read()
        s3_client = boto3.client('s3')
        file_to_save=BytesIO()
        file.save(file_to_save)
        file_to_save = file_to_save.getvalue()
        response = s3_client.put_object(Bucket=bucket_link,
                                        Body=(bytes(file_to_save)),
                                        Key="log/log_scraping.xlsx")
        response = s3_client.put_object(Bucket=bucket_link,
                                        Body=(bytes(file_to_save)),
                                        Key="log/log_scraping_backup.xlsx")
        return response
    except Exception as e:
        log("Couldn't save logfile.xlsx information in Amazon S3. Error description:" + str(e))

def change_rule(state,rule_name, arn, lambda_function_name, lambda_function_arn): # Change the rule in Events.; Used to keep lambda function running
    try:

        s3_client = boto3.client('events', region_name='us-west-2')
        response = s3_client.put_rule(Name=rule_name,
        ScheduleExpression="rate(15 minutes)",
        EventPattern="""{
  "source": ["aws.lambda"]
}""",
        State=state,
        Description='This rule is used to keep the scrapers running',
        RoleArn=arn)
        response = s3_client.put_targets(
            Rule=rule_name,
            Targets=[{'Id':lambda_function_name, 'Arn': lambda_function_arn}])
        response = s3_client.list_rules()
        return response
    except Exception as e:
        log("Couldn't create rule in CloudEvents. Error description:" + str(e))


# print(change_rule("DISABLED","teste", "arn:aws:iam::012782101632:role/Infra2-energyhub-LambdaRole-scraper", "teste_bid_1","arn:aws:lambda:us-east-1:012782101632:function:teste_bid_1"))